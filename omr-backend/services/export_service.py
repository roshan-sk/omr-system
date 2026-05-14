import io
from datetime import datetime

from flask import jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from models import StudentOMR, ScoringRule


def generate_export_excel(batch_id):

    try:

        students = (
            StudentOMR.query
            .filter_by(batch_id=batch_id)
            .order_by(StudentOMR.score.desc())
            .all()
        )

    except Exception as e:

        return jsonify({
            "error": f"DB error: {str(e)}"
        }), 500

    if not students:

        return jsonify({
            "error": "No results found for this batch"
        }), 404

    try:
        rules_by_level = {}

        all_rules = (
            ScoringRule.query
            .order_by(
                ScoringRule.level,
                ScoringRule.range_from
            )
            .all()
        )

        for rule in all_rules:

            level = (
                (rule.level or "")
                .strip()
                .lower()
                .replace(" ", "_")
            )

            rules_by_level.setdefault(level, []).append(rule)

        max_marks_by_level = {}

        for level, rules in rules_by_level.items():

            max_marks_by_level[level] = sum(
                float(rule.correct_marks)
                for rule in rules
                for _ in range(
                    rule.range_from,
                    rule.range_to + 1
                )
            )

        wb = Workbook()

        ws = wb.active

        ws.title = "OMR Results"

        green = PatternFill(
            start_color="C6EFCE",
            end_color="C6EFCE",
            fill_type="solid"
        )

        red = PatternFill(
            start_color="FFC7CE",
            end_color="FFC7CE",
            fill_type="solid"
        )

        grey = PatternFill(
            start_color="E7E6E6",
            end_color="E7E6E6",
            fill_type="solid"
        )

        headers = [
            "SL No",
            "File Name",
            "Name",
            "Centre Number",
            "Level",
            "DOB"
        ]

        for i in range(1, 41):

            headers.append(f"Q{str(i).zfill(2)}")

        headers += [
            "Correct",
            "Wrong",
            "Empty",
            "Score",
            "Percentage"
        ]

        ws.append(headers)

        header_fill = PatternFill(
            start_color="305496",
            end_color="305496",
            fill_type="solid"
        )

        header_font = Font(
            color="FFFFFF",
            bold=True
        )

        for col in range(1, len(headers) + 1):

            cell = ws.cell(row=1, column=col)

            cell.fill = header_fill

            cell.font = header_font

        for idx, s in enumerate(students, start=1):

            answers = (
                s.answers
                if isinstance(s.answers, dict)
                else {}
            )

            verify = (
                s.verify_ans
                if isinstance(s.verify_ans, dict)
                else {}
            )

            correct_count = 0
            wrong_count = 0
            empty_count = 0

            row = [
                idx,
                getattr(s, "file_name", "-") or "-",
                s.name or "-",
                s.centre_number or "-",
                s.level or "-",
                s.dob or "-"
            ]

            start_col = len(row) + 1

            for i in range(1, 41):

                q_key = f"Q{str(i).zfill(2)}"

                ans = answers.get(q_key, "-")

                v = (
                    verify.get(q_key, {})
                    if isinstance(verify.get(q_key), dict)
                    else {}
                )

                is_correct = v.get("is_correct", False)

                if ans in ["-", "Empty", None, ""]:

                    empty_count += 1

                    ans = "-"

                elif is_correct:

                    correct_count += 1

                else:

                    wrong_count += 1

                row.append(ans)

            score = s.score or 0

            level_key = (
                (s.level or "")
                .strip()
                .lower()
                .replace(" ", "_")
            )

            max_marks = max_marks_by_level.get(
                level_key,
                40.0
            )

            percentage = round(
                (score / max_marks * 100),
                2
            ) if max_marks else 0

            row += [
                correct_count,
                wrong_count,
                empty_count,
                score,
                percentage
            ]

            ws.append(row)

            for i in range(40):

                col = start_col + i

                cell = ws.cell(
                    row=idx + 1,
                    column=col
                )

                q_key = f"Q{str(i+1).zfill(2)}"

                v = (
                    verify.get(q_key, {})
                    if isinstance(verify.get(q_key), dict)
                    else {}
                )

                if cell.value in ["-", "Empty", None, ""]:

                    cell.fill = grey

                elif v.get("is_correct"):

                    cell.fill = green

                else:

                    cell.fill = red

        timestamp = datetime.now().strftime(
            "%Y%m%d_%H%M%S"
        )

        filename = f"OMR_Results_{timestamp}.xlsx"

        output = io.BytesIO()

        wb.save(output)

        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        )

    except Exception as e:

        return jsonify({
            "error": f"Export failed: {str(e)}"
        }), 500